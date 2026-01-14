from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException, Form, Depends
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt
import json
import aiofiles
from openpyxl import load_workbook
import PyPDF2
import io
import re
from emergentintegrations.llm.chat import LlmChat, UserMessage, FileContentWithMimeType

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Config
JWT_SECRET = os.environ.get('JWT_SECRET', 'insurance-crm-secret-key-2024')
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24

# Create upload directory
UPLOAD_DIR = ROOT_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)

app = FastAPI()
api_router = APIRouter(prefix="/api")
security = HTTPBearer(auto_error=False)

# =============================================================================
# MODELS
# =============================================================================

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: str

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    email: str
    name: str
    created_at: str

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: UserResponse

class CarrierCreate(BaseModel):
    name: str
    code: str
    description: Optional[str] = ""
    primary_key_fields: List[str] = []
    field_mappings: Dict[str, str] = {}
    custom_fields: List[str] = []  # User-defined custom fields
    # Parsing configuration
    header_row: Optional[int] = None
    data_start_row: Optional[int] = None
    file_type: Optional[str] = "auto"

class CarrierResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    code: str
    description: str
    primary_key_fields: List[str]
    field_mappings: Dict[str, str]
    custom_fields: List[str] = []
    header_row: Optional[int] = None
    data_start_row: Optional[int] = None
    file_type: Optional[str] = "auto"
    created_at: str
    updated_at: str

class AgentCreate(BaseModel):
    name: str
    agent_code: str
    email: Optional[str] = ""
    phone: Optional[str] = ""
    address: Optional[str] = ""
    commission_rate: Optional[float] = 0.0

class AgentResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    agent_code: str
    email: str
    phone: str
    address: str
    commission_rate: float
    total_payouts: float
    created_at: str

class UploadResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    filename: str
    carrier_id: str
    carrier_name: str
    status: str
    total_records: int
    processed_records: int
    conflict_count: int
    created_at: str

class ExtractedRecordResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    upload_id: str
    carrier_id: str
    raw_data: Dict[str, Any]
    mapped_data: Dict[str, Any]
    status: str
    conflicts: List[Dict[str, Any]]
    created_at: str

class ConflictResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    record_id: str
    upload_id: str
    conflict_type: str
    field_name: str
    current_value: Any
    new_value: Any
    status: str
    created_at: str

class PayoutResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    agent_id: str
    agent_name: str
    carrier_id: str
    carrier_name: str
    policy_number: str
    amount: float
    commission: float
    status: str
    payout_date: str
    created_at: str

class FieldMappingUpdate(BaseModel):
    field_mappings: Dict[str, str]
    primary_key_fields: List[str]

class ConflictResolution(BaseModel):
    resolution: str  # "accept_new", "keep_current", "manual"
    manual_value: Optional[Any] = None

class DashboardStats(BaseModel):
    total_carriers: int
    total_agents: int
    total_uploads: int
    pending_reviews: int
    total_conflicts: int
    total_payouts: float
    recent_uploads: List[Dict[str, Any]]
    recent_conflicts: List[Dict[str, Any]]

# =============================================================================
# AUTH HELPERS
# =============================================================================

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_token(user_id: str) -> str:
    payload = {
        "sub": user_id,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    if not credentials:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("sub")
        user = await db.users.find_one({"id": user_id}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

# =============================================================================
# AUTH ROUTES
# =============================================================================

@api_router.post("/auth/register", response_model=TokenResponse)
async def register(user_data: UserCreate):
    existing = await db.users.find_one({"email": user_data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user_id = str(uuid.uuid4())
    user = {
        "id": user_id,
        "email": user_data.email,
        "password": hash_password(user_data.password),
        "name": user_data.name,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user)
    
    token = create_token(user_id)
    return TokenResponse(
        access_token=token,
        user=UserResponse(
            id=user_id,
            email=user_data.email,
            name=user_data.name,
            created_at=user["created_at"]
        )
    )

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(credentials: UserLogin):
    user = await db.users.find_one({"email": credentials.email}, {"_id": 0})
    if not user or not verify_password(credentials.password, user["password"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    token = create_token(user["id"])
    return TokenResponse(
        access_token=token,
        user=UserResponse(
            id=user["id"],
            email=user["email"],
            name=user["name"],
            created_at=user["created_at"]
        )
    )

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(current_user: dict = Depends(get_current_user)):
    return UserResponse(**current_user)

# =============================================================================
# CARRIER ROUTES
# =============================================================================

@api_router.post("/carriers", response_model=CarrierResponse)
async def create_carrier(carrier: CarrierCreate, current_user: dict = Depends(get_current_user)):
    existing = await db.carriers.find_one({"code": carrier.code})
    if existing:
        raise HTTPException(status_code=400, detail="Carrier code already exists")
    
    carrier_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    carrier_doc = {
        "id": carrier_id,
        **carrier.model_dump(),
        "created_at": now,
        "updated_at": now
    }
    await db.carriers.insert_one(carrier_doc)
    return CarrierResponse(**carrier_doc)

@api_router.get("/carriers", response_model=List[CarrierResponse])
async def get_carriers(current_user: dict = Depends(get_current_user)):
    carriers = await db.carriers.find({}, {"_id": 0}).to_list(1000)
    return [CarrierResponse(**c) for c in carriers]

@api_router.get("/carriers/{carrier_id}", response_model=CarrierResponse)
async def get_carrier(carrier_id: str, current_user: dict = Depends(get_current_user)):
    carrier = await db.carriers.find_one({"id": carrier_id}, {"_id": 0})
    if not carrier:
        raise HTTPException(status_code=404, detail="Carrier not found")
    return CarrierResponse(**carrier)

@api_router.put("/carriers/{carrier_id}", response_model=CarrierResponse)
async def update_carrier(carrier_id: str, carrier: CarrierCreate, current_user: dict = Depends(get_current_user)):
    existing = await db.carriers.find_one({"id": carrier_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Carrier not found")
    
    update_data = carrier.model_dump()
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.carriers.update_one({"id": carrier_id}, {"$set": update_data})
    updated = await db.carriers.find_one({"id": carrier_id}, {"_id": 0})
    return CarrierResponse(**updated)

@api_router.put("/carriers/{carrier_id}/field-mappings", response_model=CarrierResponse)
async def update_field_mappings(carrier_id: str, mapping: FieldMappingUpdate, current_user: dict = Depends(get_current_user)):
    existing = await db.carriers.find_one({"id": carrier_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Carrier not found")
    
    await db.carriers.update_one(
        {"id": carrier_id},
        {"$set": {
            "field_mappings": mapping.field_mappings,
            "primary_key_fields": mapping.primary_key_fields,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    updated = await db.carriers.find_one({"id": carrier_id}, {"_id": 0})
    return CarrierResponse(**updated)

@api_router.delete("/carriers/{carrier_id}")
async def delete_carrier(carrier_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.carriers.delete_one({"id": carrier_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Carrier not found")
    return {"message": "Carrier deleted"}

# =============================================================================
# CUSTOM FIELDS ROUTES
# =============================================================================

@api_router.get("/custom-fields")
async def get_custom_fields(current_user: dict = Depends(get_current_user)):
    """Get all custom fields defined in the system"""
    fields = await db.custom_fields.find({}, {"_id": 0}).to_list(1000)
    return fields

@api_router.post("/custom-fields")
async def create_custom_field(
    field_name: str,
    field_label: str,
    field_type: str = "text",
    current_user: dict = Depends(get_current_user)
):
    """Create a new custom field"""
    # Check if field already exists
    existing = await db.custom_fields.find_one({"field_name": field_name})
    if existing:
        raise HTTPException(status_code=400, detail="Field already exists")
    
    field_doc = {
        "id": str(uuid.uuid4()),
        "field_name": field_name.lower().replace(" ", "_"),
        "field_label": field_label,
        "field_type": field_type,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": current_user["id"]
    }
    await db.custom_fields.insert_one(field_doc)
    return {"message": "Custom field created", "field": field_doc}

@api_router.delete("/custom-fields/{field_name}")
async def delete_custom_field(field_name: str, current_user: dict = Depends(get_current_user)):
    """Delete a custom field"""
    result = await db.custom_fields.delete_one({"field_name": field_name})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Field not found")
    return {"message": "Custom field deleted"}

# =============================================================================
# AGENT ROUTES
# =============================================================================

@api_router.post("/agents", response_model=AgentResponse)
async def create_agent(agent: AgentCreate, current_user: dict = Depends(get_current_user)):
    existing = await db.agents.find_one({"agent_code": agent.agent_code})
    if existing:
        raise HTTPException(status_code=400, detail="Agent code already exists")
    
    agent_id = str(uuid.uuid4())
    agent_doc = {
        "id": agent_id,
        **agent.model_dump(),
        "total_payouts": 0.0,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.agents.insert_one(agent_doc)
    return AgentResponse(**agent_doc)

@api_router.get("/agents", response_model=List[AgentResponse])
async def get_agents(current_user: dict = Depends(get_current_user)):
    agents = await db.agents.find({}, {"_id": 0}).to_list(1000)
    return [AgentResponse(**a) for a in agents]

@api_router.get("/agents/{agent_id}", response_model=AgentResponse)
async def get_agent(agent_id: str, current_user: dict = Depends(get_current_user)):
    agent = await db.agents.find_one({"id": agent_id}, {"_id": 0})
    if not agent:
        raise HTTPException(status_code=404, detail="Agent not found")
    return AgentResponse(**agent)

@api_router.put("/agents/{agent_id}", response_model=AgentResponse)
async def update_agent(agent_id: str, agent: AgentCreate, current_user: dict = Depends(get_current_user)):
    existing = await db.agents.find_one({"id": agent_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Agent not found")
    
    await db.agents.update_one({"id": agent_id}, {"$set": agent.model_dump()})
    updated = await db.agents.find_one({"id": agent_id}, {"_id": 0})
    return AgentResponse(**updated)

@api_router.delete("/agents/{agent_id}")
async def delete_agent(agent_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.agents.delete_one({"id": agent_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Agent not found")
    return {"message": "Agent deleted"}

# =============================================================================
# FILE PROCESSING HELPERS
# =============================================================================

async def extract_excel_data(file_path: str, header_row: int = None, start_row: int = None) -> List[Dict[str, Any]]:
    """Extract data from Excel file with smart header detection"""
    wb = load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    headers = []
    data = []
    
    # If header_row not specified, detect it by finding first row with multiple non-empty cells
    if header_row is None:
        for row_idx in range(1, min(20, sheet.max_row + 1)):
            row_values = [sheet.cell(row=row_idx, column=col).value for col in range(1, min(50, sheet.max_column + 1))]
            non_empty = [v for v in row_values if v is not None and str(v).strip()]
            # A header row typically has multiple text values
            if len(non_empty) >= 5:
                header_row = row_idx
                break
        if header_row is None:
            header_row = 1
    
    if start_row is None:
        start_row = header_row + 1
    
    # Extract headers
    for col in range(1, sheet.max_column + 1):
        val = sheet.cell(row=header_row, column=col).value
        headers.append(str(val).strip() if val else f"column_{col}")
    
    # Extract data rows
    for row_idx in range(start_row, sheet.max_row + 1):
        row_values = [sheet.cell(row=row_idx, column=col).value for col in range(1, sheet.max_column + 1)]
        
        # Skip empty rows
        if not any(v is not None for v in row_values):
            continue
            
        row_data = {}
        for i, val in enumerate(row_values):
            if i < len(headers):
                header = headers[i]
                # Clean up the value
                if val is not None:
                    row_data[header] = val
                else:
                    row_data[header] = None
        
        # Only add if row has meaningful data (at least 3 non-null values)
        non_null_count = sum(1 for v in row_data.values() if v is not None)
        if non_null_count >= 3:
            data.append(row_data)
    
    return data

async def extract_pdf_text(file_path: str) -> str:
    """Extract text from PDF file"""
    with open(file_path, 'rb') as f:
        reader = PyPDF2.PdfReader(f)
        text = ""
        for page in reader.pages:
            text += page.extract_text() + "\n"
    return text

async def ai_extract_from_pdf(file_path: str, carrier_mappings: Dict[str, str]) -> List[Dict[str, Any]]:
    """Use AI to extract structured data from PDF"""
    api_key = os.environ.get('EMERGENT_LLM_KEY')
    if not api_key:
        raise HTTPException(status_code=500, detail="AI extraction not configured")
    
    system_message = f"""You are a data extraction specialist for insurance documents. 
    Extract structured data from the document and return it as a JSON array.
    
    The expected fields to extract are: {json.dumps(list(carrier_mappings.values()))}
    
    Map the document fields to these standard fields:
    {json.dumps(carrier_mappings)}
    
    Return ONLY a valid JSON array of objects. Each object should have the mapped field names as keys.
    If a field is not found, use null.
    Extract ALL records/rows from the document."""
    
    chat = LlmChat(
        api_key=api_key,
        session_id=f"extraction-{uuid.uuid4()}",
        system_message=system_message
    ).with_model("gemini", "gemini-2.5-flash")
    
    file_content = FileContentWithMimeType(
        file_path=file_path,
        mime_type="application/pdf"
    )
    
    user_message = UserMessage(
        text="Extract all payout/commission records from this insurance document. Return as JSON array.",
        file_contents=[file_content]
    )
    
    response = await chat.send_message(user_message)
    
    # Parse JSON from response
    try:
        # Try to find JSON array in response
        json_match = re.search(r'\[[\s\S]*\]', response)
        if json_match:
            return json.loads(json_match.group())
        return []
    except json.JSONDecodeError:
        logging.error(f"Failed to parse AI response: {response}")
        return []

def apply_field_mappings(raw_data: Dict[str, Any], mappings: Dict[str, str]) -> Dict[str, Any]:
    """Apply field mappings to raw data"""
    mapped = {}
    for source_field, target_field in mappings.items():
        # Check for exact match first
        if source_field in raw_data:
            mapped[target_field] = raw_data[source_field]
        else:
            # Try case-insensitive match
            for key in raw_data:
                if key.lower() == source_field.lower():
                    mapped[target_field] = raw_data[key]
                    break
    
    # Add unmapped fields to raw section
    mapped["_raw"] = raw_data
    return mapped

async def detect_conflicts(new_data: Dict[str, Any], carrier_id: str, primary_keys: List[str]) -> List[Dict[str, Any]]:
    """Detect conflicts with existing records"""
    conflicts = []
    
    # Build query for primary key match
    query = {"carrier_id": carrier_id, "status": {"$in": ["validated", "pending"]}}
    pk_conditions = []
    
    for pk in primary_keys:
        if pk in new_data.get("mapped_data", {}):
            pk_conditions.append({f"mapped_data.{pk}": new_data["mapped_data"][pk]})
    
    if pk_conditions:
        query["$or"] = pk_conditions
    
    existing = await db.extracted_records.find_one(query, {"_id": 0})
    
    if existing:
        # Check for field mismatches
        for field, new_value in new_data.get("mapped_data", {}).items():
            if field.startswith("_"):
                continue
            existing_value = existing.get("mapped_data", {}).get(field)
            if existing_value is not None and new_value is not None:
                if str(existing_value) != str(new_value):
                    conflicts.append({
                        "type": "mismatch",
                        "field": field,
                        "existing_value": existing_value,
                        "new_value": new_value,
                        "existing_record_id": existing["id"]
                    })
        
        if not conflicts:
            conflicts.append({
                "type": "duplicate",
                "field": "record",
                "existing_record_id": existing["id"],
                "message": "Potential duplicate record found"
            })
    
    return conflicts

# =============================================================================
# UPLOAD ROUTES
# =============================================================================

@api_router.post("/uploads", response_model=UploadResponse)
async def create_upload(
    file: UploadFile = File(...),
    carrier_id: str = Form(...),
    current_user: dict = Depends(get_current_user)
):
    # Validate carrier
    carrier = await db.carriers.find_one({"id": carrier_id}, {"_id": 0})
    if not carrier:
        raise HTTPException(status_code=404, detail="Carrier not found")
    
    # Save file
    file_ext = Path(file.filename).suffix.lower()
    if file_ext not in ['.xlsx', '.xls', '.pdf', '.csv']:
        raise HTTPException(status_code=400, detail="Unsupported file type. Use Excel (.xlsx, .xls), PDF, or CSV")
    
    upload_id = str(uuid.uuid4())
    file_path = UPLOAD_DIR / f"{upload_id}{file_ext}"
    
    async with aiofiles.open(file_path, 'wb') as out_file:
        content = await file.read()
        await out_file.write(content)
    
    # Create upload record
    upload_doc = {
        "id": upload_id,
        "filename": file.filename,
        "file_path": str(file_path),
        "carrier_id": carrier_id,
        "carrier_name": carrier["name"],
        "status": "processing",
        "total_records": 0,
        "processed_records": 0,
        "conflict_count": 0,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "user_id": current_user["id"]
    }
    await db.uploads.insert_one(upload_doc)
    
    # Process file based on type
    try:
        extracted_data = []
        
        # Get carrier parsing configuration
        header_row = carrier.get("header_row")
        data_start_row = carrier.get("data_start_row")
        
        if file_ext in ['.xlsx', '.xls']:
            extracted_data = await extract_excel_data(str(file_path), header_row=header_row, start_row=data_start_row)
        elif file_ext == '.pdf':
            extracted_data = await ai_extract_from_pdf(str(file_path), carrier.get("field_mappings", {}))
        elif file_ext == '.csv':
            # Handle CSV files
            import csv
            with open(file_path, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                extracted_data = list(reader)
        
        # Process each record
        conflict_count = 0
        for idx, raw_record in enumerate(extracted_data):
            mapped_data = apply_field_mappings(raw_record, carrier.get("field_mappings", {}))
            
            record_doc = {
                "id": str(uuid.uuid4()),
                "upload_id": upload_id,
                "carrier_id": carrier_id,
                "raw_data": raw_record,
                "mapped_data": mapped_data,
                "status": "pending",
                "conflicts": [],
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            
            # Check for conflicts
            conflicts = await detect_conflicts(record_doc, carrier_id, carrier.get("primary_key_fields", []))
            if conflicts:
                record_doc["conflicts"] = conflicts
                record_doc["status"] = "conflict"
                conflict_count += len(conflicts)
                
                # Create conflict records
                for conflict in conflicts:
                    conflict_doc = {
                        "id": str(uuid.uuid4()),
                        "record_id": record_doc["id"],
                        "upload_id": upload_id,
                        "existing_record_id": conflict.get("existing_record_id"),
                        "conflict_type": conflict["type"],
                        "field_name": conflict.get("field", ""),
                        "current_value": conflict.get("existing_value"),
                        "new_value": conflict.get("new_value"),
                        "status": "pending",
                        "created_at": datetime.now(timezone.utc).isoformat()
                    }
                    await db.conflicts.insert_one(conflict_doc)
            
            await db.extracted_records.insert_one(record_doc)
        
        # Update upload status
        await db.uploads.update_one(
            {"id": upload_id},
            {"$set": {
                "status": "completed",
                "total_records": len(extracted_data),
                "processed_records": len(extracted_data),
                "conflict_count": conflict_count
            }}
        )
        
        upload_doc["status"] = "completed"
        upload_doc["total_records"] = len(extracted_data)
        upload_doc["processed_records"] = len(extracted_data)
        upload_doc["conflict_count"] = conflict_count
        
    except Exception as e:
        logging.error(f"Error processing upload: {e}")
        await db.uploads.update_one(
            {"id": upload_id},
            {"$set": {"status": "error", "error": str(e)}}
        )
        upload_doc["status"] = "error"
    
    return UploadResponse(**upload_doc)

@api_router.get("/uploads", response_model=List[UploadResponse])
async def get_uploads(current_user: dict = Depends(get_current_user)):
    uploads = await db.uploads.find({}, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return [UploadResponse(**u) for u in uploads]

@api_router.get("/uploads/{upload_id}", response_model=UploadResponse)
async def get_upload(upload_id: str, current_user: dict = Depends(get_current_user)):
    upload = await db.uploads.find_one({"id": upload_id}, {"_id": 0})
    if not upload:
        raise HTTPException(status_code=404, detail="Upload not found")
    return UploadResponse(**upload)

@api_router.delete("/uploads/{upload_id}")
async def delete_upload(upload_id: str, current_user: dict = Depends(get_current_user)):
    """Delete an upload and all its associated records and conflicts"""
    upload = await db.uploads.find_one({"id": upload_id})
    if not upload:
        raise HTTPException(status_code=404, detail="Upload not found")
    
    # Delete all conflicts for records in this upload
    await db.conflicts.delete_many({"upload_id": upload_id})
    
    # Delete all records for this upload
    deleted_records = await db.extracted_records.delete_many({"upload_id": upload_id})
    
    # Delete the upload itself
    await db.uploads.delete_one({"id": upload_id})
    
    # Delete the file if it exists
    file_path = Path(upload.get("file_path", ""))
    if file_path.exists():
        try:
            file_path.unlink()
        except:
            pass
    
    return {
        "message": "Upload deleted successfully",
        "deleted_records": deleted_records.deleted_count
    }

@api_router.get("/uploads/{upload_id}/records", response_model=List[ExtractedRecordResponse])
async def get_upload_records(upload_id: str, current_user: dict = Depends(get_current_user)):
    records = await db.extracted_records.find({"upload_id": upload_id}, {"_id": 0}).to_list(10000)
    return [ExtractedRecordResponse(**r) for r in records]

@api_router.post("/uploads/preview")
async def preview_file_structure(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Preview Excel file structure to help configure carrier parsing settings"""
    file_ext = Path(file.filename).suffix.lower()
    if file_ext not in ['.xlsx', '.xls']:
        raise HTTPException(status_code=400, detail="Preview only supports Excel files")
    
    # Save temp file
    temp_path = UPLOAD_DIR / f"temp_preview_{uuid.uuid4()}{file_ext}"
    async with aiofiles.open(temp_path, 'wb') as out_file:
        content = await file.read()
        await out_file.write(content)
    
    try:
        wb = load_workbook(str(temp_path), data_only=True)
        sheet = wb.active
        
        # Get first 10 rows for preview
        preview_rows = []
        for row_idx in range(1, min(11, sheet.max_row + 1)):
            row_data = []
            for col in range(1, min(20, sheet.max_column + 1)):
                val = sheet.cell(row=row_idx, column=col).value
                row_data.append(str(val)[:50] if val else None)
            preview_rows.append({
                "row_number": row_idx,
                "values": row_data
            })
        
        # Detect likely header row
        suggested_header_row = 1
        for row in preview_rows:
            non_null_count = sum(1 for v in row["values"] if v is not None)
            if non_null_count >= 5:
                suggested_header_row = row["row_number"]
                break
        
        return {
            "filename": file.filename,
            "total_rows": sheet.max_row,
            "total_columns": sheet.max_column,
            "preview_rows": preview_rows,
            "suggested_header_row": suggested_header_row,
            "suggested_data_start_row": suggested_header_row + 1
        }
    finally:
        if temp_path.exists():
            temp_path.unlink()

# =============================================================================
# RECORDS & CONFLICTS ROUTES
# =============================================================================

@api_router.get("/records", response_model=List[ExtractedRecordResponse])
async def get_records(
    status: Optional[str] = None,
    carrier_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if status:
        query["status"] = status
    if carrier_id:
        query["carrier_id"] = carrier_id
    
    records = await db.extracted_records.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return [ExtractedRecordResponse(**r) for r in records]

@api_router.put("/records/{record_id}/validate")
async def validate_record(record_id: str, current_user: dict = Depends(get_current_user)):
    record = await db.extracted_records.find_one({"id": record_id})
    if not record:
        raise HTTPException(status_code=404, detail="Record not found")
    
    await db.extracted_records.update_one(
        {"id": record_id},
        {"$set": {"status": "validated"}}
    )
    
    # Resolve any pending conflicts for this record
    await db.conflicts.update_many(
        {"record_id": record_id},
        {"$set": {"status": "resolved"}}
    )
    
    return {"message": "Record validated"}

@api_router.put("/records/{record_id}/reject")
async def reject_record(record_id: str, current_user: dict = Depends(get_current_user)):
    await db.extracted_records.update_one(
        {"id": record_id},
        {"$set": {"status": "rejected"}}
    )
    
    await db.conflicts.update_many(
        {"record_id": record_id},
        {"$set": {"status": "rejected"}}
    )
    
    return {"message": "Record rejected"}

@api_router.get("/conflicts", response_model=List[ConflictResponse])
async def get_conflicts(
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if status:
        query["status"] = status
    
    conflicts = await db.conflicts.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return [ConflictResponse(**c) for c in conflicts]

@api_router.get("/conflicts/{conflict_id}/details")
async def get_conflict_details(conflict_id: str, current_user: dict = Depends(get_current_user)):
    """Get detailed conflict info including source upload information"""
    conflict = await db.conflicts.find_one({"id": conflict_id}, {"_id": 0})
    if not conflict:
        raise HTTPException(status_code=404, detail="Conflict not found")
    
    # Get the new record (from current upload)
    new_record = await db.extracted_records.find_one({"id": conflict["record_id"]}, {"_id": 0})
    new_upload = None
    if new_record:
        new_upload = await db.uploads.find_one({"id": new_record.get("upload_id")}, {"_id": 0})
    
    # Get the existing record that caused the conflict
    existing_record = None
    existing_upload = None
    if conflict.get("existing_record_id"):
        existing_record = await db.extracted_records.find_one({"id": conflict["existing_record_id"]}, {"_id": 0})
        if existing_record:
            existing_upload = await db.uploads.find_one({"id": existing_record.get("upload_id")}, {"_id": 0})
    
    return {
        "conflict": conflict,
        "new_record": new_record,
        "new_upload": {
            "filename": new_upload.get("filename") if new_upload else None,
            "carrier_name": new_upload.get("carrier_name") if new_upload else None,
            "created_at": new_upload.get("created_at") if new_upload else None
        } if new_upload else None,
        "existing_record": existing_record,
        "existing_upload": {
            "filename": existing_upload.get("filename") if existing_upload else None,
            "carrier_name": existing_upload.get("carrier_name") if existing_upload else None,
            "created_at": existing_upload.get("created_at") if existing_upload else None
        } if existing_upload else None,
        "reason": _get_conflict_reason(conflict)
    }

def _get_conflict_reason(conflict: dict) -> str:
    """Generate human-readable conflict reason"""
    ctype = conflict.get("conflict_type", "unknown")
    field = conflict.get("field_name", "unknown field")
    
    if ctype == "mismatch":
        return f"The field '{field}' has different values in the existing record vs the new upload. This happens when the same record (matched by primary key) has conflicting data."
    elif ctype == "duplicate":
        return "A record with the same primary key values already exists. This could be a true duplicate or an update to existing data."
    else:
        return "Data conflict detected between existing and newly uploaded records."

@api_router.get("/export/approved")
async def export_approved_data(
    format: str = "csv",
    carrier_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Export approved data in various formats"""
    query = {"status": "validated"}
    if carrier_id:
        query["carrier_id"] = carrier_id
    
    records = await db.extracted_records.find(query, {"_id": 0}).to_list(100000)
    
    if not records:
        raise HTTPException(status_code=404, detail="No approved data to export")
    
    # Get all unique fields
    all_fields = set()
    for r in records:
        for k in r.get("mapped_data", {}).keys():
            if not k.startswith("_"):
                all_fields.add(k)
    fields = sorted(list(all_fields))
    
    if format == "json":
        # JSON format - good for API integrations
        export_data = []
        for r in records:
            row = {"_record_id": r["id"]}
            for f in fields:
                row[f] = r.get("mapped_data", {}).get(f)
            export_data.append(row)
        return {"format": "json", "count": len(export_data), "data": export_data}
    
    elif format == "zoho":
        # Zoho CRM format - maps to Zoho Account and Contact fields
        export_data = []
        for r in records:
            mapped = r.get("mapped_data", {})
            row = {
                # Account fields (from Zoho Accounts)
                "AC_Account_Number": mapped.get("broker_id") or mapped.get("agent_code", ""),
                "Account_Name": mapped.get("broker_name") or mapped.get("agent_name", ""),
                "Account_Owner": "",  # Would need to be set in Zoho
                "Account_Type": "Brokers - Member",  # Default type
                "Entity_Type": "LLC",  # Default, would need mapping
                
                # Contact fields (from Zoho Contacts)
                "Primary_Contact": mapped.get("agent_name", ""),
                "Email": mapped.get("email", ""),
                "Mobile": mapped.get("phone", ""),
                "Mailing_Street": mapped.get("address", ""),
                "Mailing_City": mapped.get("city", ""),
                "Mailing_State": mapped.get("state", ""),
                "Mailing_Zip": mapped.get("zip", ""),
                
                # Business metrics
                "New_WP": mapped.get("new_wp") or mapped.get("new_written_premium", ""),
                "Total_WP": mapped.get("total_wp") or mapped.get("premium") or mapped.get("amount", ""),
                "Earned_Premium": mapped.get("earned", ""),
                "Incurred": mapped.get("incurred", ""),
                "Loss_Ratio": mapped.get("loss_ratio", ""),
                "Policy_Type": mapped.get("policy_type", ""),
                "PG_Code": mapped.get("pg_code", ""),
                
                # Metadata
                "Record_Source": "Rev-Box Import",
                "Import_Date": datetime.now(timezone.utc).strftime("%Y-%m-%d")
            }
            export_data.append(row)
        return {"format": "zoho", "count": len(export_data), "data": export_data}
    
    else:
        # CSV format (default)
        import io
        import csv
        
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=fields)
        writer.writeheader()
        
        for r in records:
            row = {}
            for f in fields:
                val = r.get("mapped_data", {}).get(f)
                row[f] = str(val) if val is not None else ""
            writer.writerow(row)
        
        return {
            "format": "csv",
            "count": len(records),
            "csv_content": output.getvalue()
        }

@api_router.put("/conflicts/{conflict_id}/resolve")
async def resolve_conflict(
    conflict_id: str,
    resolution: ConflictResolution,
    current_user: dict = Depends(get_current_user)
):
    conflict = await db.conflicts.find_one({"id": conflict_id}, {"_id": 0})
    if not conflict:
        raise HTTPException(status_code=404, detail="Conflict not found")
    
    # Update conflict status
    await db.conflicts.update_one(
        {"id": conflict_id},
        {"$set": {"status": "resolved", "resolution": resolution.resolution}}
    )
    
    # If accepting new value, update the record
    if resolution.resolution == "accept_new":
        record = await db.extracted_records.find_one({"id": conflict["record_id"]})
        if record and conflict["field_name"]:
            mapped_data = record.get("mapped_data", {})
            mapped_data[conflict["field_name"]] = conflict["new_value"]
            await db.extracted_records.update_one(
                {"id": conflict["record_id"]},
                {"$set": {"mapped_data": mapped_data}}
            )
    elif resolution.resolution == "manual" and resolution.manual_value is not None:
        record = await db.extracted_records.find_one({"id": conflict["record_id"]})
        if record and conflict["field_name"]:
            mapped_data = record.get("mapped_data", {})
            mapped_data[conflict["field_name"]] = resolution.manual_value
            await db.extracted_records.update_one(
                {"id": conflict["record_id"]},
                {"$set": {"mapped_data": mapped_data}}
            )
    
    # Check if all conflicts for this record are resolved
    remaining = await db.conflicts.count_documents({
        "record_id": conflict["record_id"],
        "status": "pending"
    })
    
    if remaining == 0:
        await db.extracted_records.update_one(
            {"id": conflict["record_id"]},
            {"$set": {"status": "pending"}}  # Ready for validation
        )
    
    return {"message": "Conflict resolved"}

# =============================================================================
# PAYOUT ROUTES
# =============================================================================

@api_router.post("/payouts/generate")
async def generate_payouts(
    record_ids: List[str],
    current_user: dict = Depends(get_current_user)
):
    """Generate payout records from validated extracted records"""
    payouts_created = 0
    
    for record_id in record_ids:
        record = await db.extracted_records.find_one({"id": record_id, "status": "validated"}, {"_id": 0})
        if not record:
            continue
        
        mapped = record.get("mapped_data", {})
        carrier = await db.carriers.find_one({"id": record["carrier_id"]}, {"_id": 0})
        
        # Try to find matching agent
        agent_code = mapped.get("agent_code") or mapped.get("agent_id") or mapped.get("producer_code")
        agent = None
        if agent_code:
            agent = await db.agents.find_one({"agent_code": str(agent_code)}, {"_id": 0})
        
        amount = float(mapped.get("amount", 0) or mapped.get("premium", 0) or mapped.get("payout_amount", 0) or 0)
        commission_rate = agent.get("commission_rate", 0) if agent else 0
        commission = amount * (commission_rate / 100)
        
        payout_doc = {
            "id": str(uuid.uuid4()),
            "record_id": record_id,
            "agent_id": agent["id"] if agent else None,
            "agent_name": agent["name"] if agent else "Unknown",
            "carrier_id": record["carrier_id"],
            "carrier_name": carrier["name"] if carrier else "Unknown",
            "policy_number": str(mapped.get("policy_number", "") or mapped.get("policy_id", "") or ""),
            "amount": amount,
            "commission": commission,
            "status": "pending",
            "payout_date": datetime.now(timezone.utc).isoformat(),
            "created_at": datetime.now(timezone.utc).isoformat(),
            "mapped_data": mapped
        }
        
        await db.payouts.insert_one(payout_doc)
        payouts_created += 1
        
        # Update agent total
        if agent:
            await db.agents.update_one(
                {"id": agent["id"]},
                {"$inc": {"total_payouts": amount}}
            )
        
        # Mark record as processed
        await db.extracted_records.update_one(
            {"id": record_id},
            {"$set": {"status": "processed"}}
        )
    
    return {"message": f"Created {payouts_created} payout records"}

@api_router.get("/payouts", response_model=List[PayoutResponse])
async def get_payouts(
    status: Optional[str] = None,
    agent_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if status:
        query["status"] = status
    if agent_id:
        query["agent_id"] = agent_id
    
    payouts = await db.payouts.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return [PayoutResponse(**p) for p in payouts]

@api_router.put("/payouts/{payout_id}/complete")
async def complete_payout(payout_id: str, current_user: dict = Depends(get_current_user)):
    result = await db.payouts.update_one(
        {"id": payout_id},
        {"$set": {"status": "completed"}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Payout not found")
    return {"message": "Payout marked as completed"}

# =============================================================================
# DASHBOARD ROUTE
# =============================================================================

@api_router.get("/dashboard", response_model=DashboardStats)
async def get_dashboard(current_user: dict = Depends(get_current_user)):
    total_carriers = await db.carriers.count_documents({})
    total_agents = await db.agents.count_documents({})
    total_uploads = await db.uploads.count_documents({})
    pending_reviews = await db.extracted_records.count_documents({"status": {"$in": ["pending", "conflict"]}})
    total_conflicts = await db.conflicts.count_documents({"status": "pending"})
    
    # Calculate total payouts
    pipeline = [{"$group": {"_id": None, "total": {"$sum": "$amount"}}}]
    payout_result = await db.payouts.aggregate(pipeline).to_list(1)
    total_payouts = payout_result[0]["total"] if payout_result else 0
    
    # Recent uploads
    recent_uploads = await db.uploads.find({}, {"_id": 0}).sort("created_at", -1).to_list(5)
    
    # Recent conflicts
    recent_conflicts = await db.conflicts.find({"status": "pending"}, {"_id": 0}).sort("created_at", -1).to_list(5)
    
    return DashboardStats(
        total_carriers=total_carriers,
        total_agents=total_agents,
        total_uploads=total_uploads,
        pending_reviews=pending_reviews,
        total_conflicts=total_conflicts,
        total_payouts=total_payouts,
        recent_uploads=recent_uploads,
        recent_conflicts=recent_conflicts
    )

# =============================================================================
# AI FIELD SUGGESTION
# =============================================================================

@api_router.post("/carriers/{carrier_id}/suggest-mappings")
async def suggest_field_mappings(
    carrier_id: str,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Use AI to suggest field mappings from a sample file"""
    carrier = await db.carriers.find_one({"id": carrier_id}, {"_id": 0})
    if not carrier:
        raise HTTPException(status_code=404, detail="Carrier not found")
    
    api_key = os.environ.get('EMERGENT_LLM_KEY')
    if not api_key:
        raise HTTPException(status_code=500, detail="AI not configured")
    
    # Save temp file
    file_ext = Path(file.filename).suffix.lower()
    temp_path = UPLOAD_DIR / f"temp_{uuid.uuid4()}{file_ext}"
    
    async with aiofiles.open(temp_path, 'wb') as out_file:
        content = await file.read()
        await out_file.write(content)
    
    try:
        # Get sample data
        if file_ext in ['.xlsx', '.xls']:
            sample_data = await extract_excel_data(str(temp_path))
            sample_fields = list(sample_data[0].keys()) if sample_data else []
        else:
            # For PDF, use AI to identify fields
            sample_fields = []
        
        system_message = """You are an expert at mapping insurance document fields to standard CRM fields.
        Given a list of source fields from an insurance carrier document, suggest mappings to these standard fields:
        - policy_number: The policy identifier
        - agent_code: The agent/producer identifier
        - agent_name: The agent's name
        - amount: The payout/premium amount
        - commission: Commission amount
        - effective_date: Policy effective date
        - insured_name: Name of the insured party
        - carrier_name: Insurance carrier name
        
        Return a JSON object with source fields as keys and suggested standard fields as values.
        Also identify which fields should be used as primary keys for duplicate detection.
        
        Format:
        {
            "mappings": {"source_field": "standard_field", ...},
            "primary_keys": ["field1", "field2"]
        }"""
        
        chat = LlmChat(
            api_key=api_key,
            session_id=f"mapping-{uuid.uuid4()}",
            system_message=system_message
        ).with_model("gemini", "gemini-2.5-flash")
        
        if file_ext == '.pdf':
            file_content = FileContentWithMimeType(
                file_path=str(temp_path),
                mime_type="application/pdf"
            )
            user_message = UserMessage(
                text="Analyze this insurance document and suggest field mappings to standard CRM fields.",
                file_contents=[file_content]
            )
        else:
            user_message = UserMessage(
                text=f"These are the fields from an insurance carrier document: {json.dumps(sample_fields)}\n\nSuggest mappings to standard CRM fields."
            )
        
        response = await chat.send_message(user_message)
        
        # Parse response
        json_match = re.search(r'\{[\s\S]*\}', response)
        if json_match:
            suggestions = json.loads(json_match.group())
            return suggestions
        
        return {"mappings": {}, "primary_keys": [], "raw_response": response}
        
    finally:
        # Clean up temp file
        if temp_path.exists():
            temp_path.unlink()

# =============================================================================
# SETUP
# =============================================================================

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
